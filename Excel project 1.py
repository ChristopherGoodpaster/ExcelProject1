import openpyxl
import os
from datetime import datetime

# Load or create Excel file
file_name = "weekdays.xlsx"
if not os.path.exists(file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Set headers for days of the week
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    for index, day in enumerate(days, start=1):
        sheet.cell(row=index, column=1).value = day
    workbook.save(file_name)

def update_day_count():
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    today = datetime.today().weekday()  # Monday = 0, Sunday = 6
    # Get the current value, if None, start from 0
    current_value = sheet.cell(row=today+1, column=2).value or 0
    sheet.cell(row=today+1, column=2).value = current_value + 1
    workbook.save(file_name)
    print(f"Updated {sheet.cell(row=today+1, column=1).value}: {current_value + 1}")

# Call this function when the Stream Deck button is pressed
update_day_count()



from Stream Deck.DeviceManager import DeviceManager
from StreamDeck.Devices import StreamDeck

def key_change(deck, key, state):
    if state:  # When the button is pressed down
        update_day_count()

# Discover Stream Deck devices
decks = DeviceManager().enumerate()
if len(decks) == 0:
    print("No Stream Deck devices found.")
else:
    deck = decks[0]  # Select the first device
    deck.open()
    deck.set_brightness(30)
    deck.set_key_callback(key_change)

    # Keep the program running
    try:
        while True:
            pass
    except KeyboardInterrupt:
        deck.close()

#End
